# https://youtu.be/7YS6YDQKFh0 : video for help 
# https://openpyxl.readthedocs.io/en/stable/styles.html : documentation 
import openpyxl
from openpyxl import Workbook , load_workbook
wb = load_workbook('index.xlsx')
# print(wb)
# sheets from workbook
ws = wb.active
print(ws)
# #-----access cell value from sheet-----
print(ws['A1'])
print(ws['A1'].value)
print(ws['B1'].value)
# #-----changing cell value -----
print(ws['B2'].value)
ws['B2'].value = "Aditi Mahabole"
print(ws['B2'].value)
# #-----to save the value-----
wb.save('index.xlsx')
# #-----accessing all sheets-----
print(wb.sheetnames)
# #----create sheet----
wb.create_sheet("Test")
print(wb.sheetnames)
# --------------------------------------------------
from openpyxl import Workbook , load_workbook
wb = Workbook()
ws = wb.active
ws.title = "Data"
#-----insert a lot of data  in sheet-----
# ws.append(["Bubu","is","great","!"])
# ws.append(["Bububui","is","great","!"])
#----loop thru----
wb.save('Test.xlsx')
# ------------------------------------------------
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('index.xlsx')
ws = wb.active
for row in range(10,15):
    for col in range(1,4):
    # for col in range(0,4):
        # char = chr(65 + col)
        char = get_column_letter(col)
        # it gives the A , B , C
        # print(ws[char+str(row)],end="  ")
        ws[char+str(row)].value = char+str(row)
        print(ws[char+str(row)] , end = " , ")
    print("\n")
wb.save('index.xlsx')
# ---output----
# 1  Aditi Mahabole  20
# =A2+1  Molshree  21
# =A3+1  Omkar  18
# =A4+1  Arushi  20
# =A5+1  Sam  19
# =A6+1  Muskan  21
# =A7+1  Manika  18
# =A8+1  Pragya  18
# =A9+1  Shivam  22
# =A10+1  Vridhi  22
# None  None  None
# None  None  None
# None  None  None
# -----------------------------------------------
# ---Merge cells----
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('index.xlsx')
ws = wb.active
ws.merge_cells("A1:C1")
ws.merge_cells("A2:C2")
ws.merge_cells("D2:G2")
ws.unmerge_cells("A1:C1")
wb.save('index.xlsx')
# -----------------------------------------------
# ---Insert Empty Rows ---
from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('index.xlsx')
ws = wb.active
# ws.insert_rows(7)
# ws.insert_rows(7)
# ws.insert_rows(7)
# ws.delete_rows(7)
# ws.delete_rows(7)
# ws.delete_rows(7)
# ws.insert_cols(1)
wb.save('index.xlsx')
# ---------------------------------------------
# ----Shift cells to other place-----
from openpyxl import Workbook , load_workbook
wb = load_workbook('index.xlsx')
ws = wb.active
ws.move_range("B3:D5" , rows = 3, cols = 3)
wb.save('index.xlsx')
# -------------------------------------------------
# -----Making a new File----
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font , PatternFill
data = {
    "Joe":{
    "math":65,
    "science":45,
    "english":67,
    "gym":89
    },
    "bubu":{
    "math":89,
    "science":95,
    "english":37,
    "gym":89
    },
    "Buddy":{
    "math":69,
    "science":23,
    "english":79,
    "gym":80
    },
    "Bubuiya":{
    "math":100,
    "science":99,
    "english":98,
    "gym":97
    },
}
wb = Workbook()
ws = wb.active
ws.title = "Grades"
headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)
# ----calculating Avg---
# print(len(['Joe']) + 3)
for col in range(2,len(['Joe']) + 5):
    char = get_column_letter(col)
    ws[char + "6"] = f"=SUM({char + '2'}:{char + '5'})/{len(data)}"
    print(ws[char + "6"].value)
# B7,C7,D7,E7   sum of B2 -->B6 , C2-->C6
for col in range(1,6):
    ws[get_column_letter(col) + '1'].font = Font(bold = True,color ="0099CCFF")
wb.save("NewGrades.xlsx")


